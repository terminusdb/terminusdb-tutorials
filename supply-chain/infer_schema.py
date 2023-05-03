#!/usr/bin/env python3
from terminusdb_client import Client
import json
import urllib.parse
import openpyxl
import itertools
import re

# Give the location of the file
path = "data/SupplyChainLogisticsProblem.xlsx"
MAX_ROW_SEARCH=1
OID_REGEXP=r"^(?P<Port>PORT\d*)|(?P<Plant>(PLANT|CND)\d*)|(?P<Customer>V(\d|_)*)$"

def type_map(cell):
    value = cell.value
    ty = cell.data_type
    if ty == 'n':
        return "xsd:decimal"
    elif ty == 's':
        matches = re.match(OID_REGEXP, value)
        if matches:
            matchdict = matches.groupdict()
            for key in matchdict:
                if matchdict[key]:
                    return key
            None
        else:
            return "xsd:string"
    elif ty == 'd':
        return "xsd:dateTime"

def infer_type(sheet_obj, j):
    ty = None
    for i in range(2, 2+MAX_ROW_SEARCH):
        val = sheet_obj.cell(row = i, column = j).value
        if val == None:
            break
        else:
            ty = type_map(sheet_obj.cell(row = i, column = j))
    return ty

def infer_schema(wb_obj):
    schema = []
    for sheet_name in wb_obj.sheetnames:
        schema_obj = { '@type' : 'Class', '@id' : sheet_name }
        object_type = sheet_name
        sheet_obj = wb_obj[sheet_name]
        # Collect header
        for i in itertools.count(start=1):
            cell = sheet_obj.cell(row = 1, column = i)
            property_name = cell.value
            if property_name == None:
                break
            else:
                schema_obj[property_name] = infer_type(sheet_obj, i)

        schema.append(schema_obj)
    return schema

if __name__ == "__main__":
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    prefixes = [{'@type' : '@context',
                 '@base' : 'http://lib.terminusdb.com/SupplyChain/',
                 '@schema' : 'http://lib.terminusdb.com/SupplyChain#'}]
    schema = infer_schema(wb_obj)
    auxiliary = [ { "@type" : "Class", "@id" : "Port", "id" : "xsd:string" },
                  { "@type" : "Class", "@id" : "Plant", "id" : "xsd:string" },
                  { "@type" : "Class", "@id" : "Customer", "id" : "xsd:string" } ]
    with open('supply_chain.json', 'w') as f:
        f.write(json.dumps(prefixes+schema+auxiliary))
