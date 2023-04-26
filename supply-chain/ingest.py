#!/usr/bin/env python3
from terminusdb_client import Client
import re
import json
import urllib.parse
import openpyxl
import itertools
import infer_schema
import datetime

team = "admin" # os.environ['TERMINUSDB_TEAM']
team_quoted = urllib.parse.quote(team)

#client = Client(f"https://cloud.terminusdb.com/{team_quoted}/")
client = Client("http://127.0.0.1:6363")
# make sure you have put the token in environment variable
# https://terminusdb.com/docs/terminuscms/get-started
client.connect(team=team)
dbid = "supply_chain"
label = "Supply Chain"
description = "A knowledge graph of supply chain information."

# Give the location of the file
path = "data/SupplyChainLogisticsProblem.xlsx"

def base_type(ty):
    return ty[0:3] == 'xsd'

def import_schema(client):
    with open('supply_chain.json', 'r') as f:
        schema = json.load(f)
    client.insert_document(schema, graph_type='schema', compress='never', full_replace=True)
    return schema

def generate_schema_dict(schema):
    schema_dict = {}
    for sobj in schema:
        if sobj['@type']== '@context':
            key = '@context'
        else:
            key = sobj['@id']
        schema_dict[key] = sobj

    return schema_dict

def import_data(client,schema):
    objects = []
    stubs = {}
    wb_obj = openpyxl.load_workbook(path)
    for sheet_name in wb_obj.sheetnames:
        print(sheet_name)
        object_type = sheet_name
        sheet_obj = wb_obj[sheet_name]
        cls = schema[object_type]
        property_types = []
        # Collect header
        for i in itertools.count(start=1):
            property_type = sheet_obj.cell(row = 1, column = i).value
            if property_type == None:
                break
            else:
                property_types.append(property_type)
        for j in itertools.count(start=2):
            obj = { '@type' : object_type }
            value = None
            for i, property_type in enumerate(property_types, 1):
                #print(f"{i} {j}")
                value_cell = sheet_obj.cell(row = j, column = i)
                value = value_cell.value
                if value == None:
                    break
                else:
                    rng = cls[property_type]
                    #print(rng)
                    #print(f"property_type: {property_type} range: {rng} value: {value}")
                    if base_type(rng):
                        if 'xsd:dateTime' == rng:
                            obj[property_type] = value.isoformat()
                        else:
                            obj[property_type] = value
                    else:
                        matches = re.match(infer_schema.OID_REGEXP, value)
                        if matches:
                            matchdict = matches.groupdict()
                            for key in matchdict:
                                if matchdict[key]:
                                    stubs[value] = key
                            obj[property_type] = { '@ref' : value }
            if value == None:
                break
            objects.append(obj)
    for oid in stubs:
        object_type = stubs[oid]
        objects.append({'@type' : object_type, '@capture' : oid })
    with open('objects.json', 'w') as f:
        f.write(json.dumps(objects))
    client.insert_document(objects, compress='never')

if __name__ == "__main__":
    exists = client.has_database(dbid)

    if exists:
        client.delete_database(dbid, team=team, force=True)

    client.create_database(dbid,
                           team,
                           label=label,
                           description=description)
    schema = import_schema(client)
    import_data(client, generate_schema_dict(schema))

